import logging
import os
import uuid
from typing import List, Dict, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.config import settings

logger = logging.getLogger(__name__)


async def create_spreadsheet(
    artifact_id: str,
    title: str,
    data: List[Dict[str, Any]],
    headers: List[str] = None,
) -> str:
    """
    Create an Excel spreadsheet.

    Args:
        artifact_id: Unique identifier
        title: Sheet title
        data: List of row dictionaries
        headers: Optional custom headers

    Returns:
        Path to the created spreadsheet
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        if data:
            if headers is None:
                headers = list(data[0].keys()) if isinstance(data[0], dict) else [f"Column {i+1}" for i in range(len(data[0]))]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(data, 2):
            if isinstance(row_data, dict):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        os.makedirs(settings.ARTIFACT_DIR, exist_ok=True)
        filename = f"{artifact_id}.xlsx"
        filepath = os.path.join(settings.ARTIFACT_DIR, filename)

        wb.save(filepath)
        logger.info(f"Created spreadsheet: {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"Spreadsheet creation error: {e}")
        raise


async def create_analysis_spreadsheet(
    artifact_id: str,
    results: Dict[str, Any],
    input_filename: str = "data.csv",
) -> str:
    """
    Create a spreadsheet for data analysis results.
    """
    try:
        wb = Workbook()

        summary_ws = wb.active
        summary_ws.title = "Summary"

        summary_data = [
            ["Analysis Summary"],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Input File", input_filename],
            ["Artifact ID", artifact_id],
            [],
        ]

        for row_idx, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_idx, column=1, value=label)
            if isinstance(value, str):
                summary_ws.cell(row=row_idx, column=2, value=value)

        if "results" in results:
            results_ws = wb.create_sheet("Results")
            results_data = results["results"]

            if results_data and isinstance(results_data[0], dict):
                headers = list(results_data[0].keys())
                for col, header in enumerate(headers, 1):
                    results_ws.cell(row=1, column=col, value=header)

                for row_idx, row in enumerate(results_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        results_ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

        os.makedirs(settings.ARTIFACT_DIR, exist_ok=True)
        filename = f"{artifact_id}.xlsx"
        filepath = os.path.join(settings.ARTIFACT_DIR, filename)

        wb.save(filepath)
        return filepath

    except Exception as e:
        logger.error(f"Analysis spreadsheet creation error: {e}")
        raise
