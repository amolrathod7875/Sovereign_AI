import logging
from typing import Dict, Any, List, Optional
import uuid
import hashlib

logger = logging.getLogger(__name__)


async def ingest_document_pipeline(document_id: str) -> Dict[str, Any]:
    """
    Complete document ingestion pipeline:
    1. Parse document (PyMuPDF for PDF, native for DOCX/XLSX)
    2. Detect if OCR is needed (scanned pages)
    3. OCR if needed (PaddleOCR)
    4. Structure-aware chunking
    5. Generate dense embeddings
    6. Generate BM25 tokens
    7. Index in Qdrant

    Args:
        document_id: ID of the uploaded document

    Returns:
        Ingestion result with stats
    """
    from app.storage.postgres import get_document_by_id

    doc = await get_document_by_id(document_id)
    if not doc:
        raise ValueError(f"Document {document_id} not found")

    logger.info(f"Starting ingestion for document: {doc.filename}")

    stages = {
        "parsing": 0,
        "ocr": 0,
        "chunking": 0,
        "embedding": 0,
        "indexing": 0,
    }

    try:
        stages["parsing"] = 1
        text_content = await parse_document(document_id, doc.filename)

        stages["parsing"] = 2
        logger.info(f"Parsed document, extracted {len(text_content)} characters")

        needs_ocr = await detect_scanned_pages(document_id)
        if needs_ocr:
            stages["ocr"] = 1
            text_content = await perform_ocr(document_id)
            stages["ocr"] = 2
            logger.info("OCR completed")

        stages["chunking"] = 1
        chunks = await chunk_text(text_content, doc)
        stages["chunking"] = 2
        logger.info(f"Created {len(chunks)} chunks")

        stages["embedding"] = 1
        await generate_embeddings(document_id, chunks)
        stages["embedding"] = 2
        logger.info("Embeddings generated")

        stages["indexing"] = 1
        await index_chunks(document_id, chunks)
        stages["indexing"] = 2
        logger.info("Chunks indexed in Qdrant")

        return {
            "document_id": document_id,
            "status": "completed",
            "stages": stages,
            "chunks": len(chunks),
            "characters": len(text_content),
        }

    except Exception as e:
        logger.error(f"Ingestion error: {e}")
        return {
            "document_id": document_id,
            "status": "failed",
            "error": str(e),
            "stages": stages,
        }


async def parse_document(document_id: str, filename: str) -> str:
    """
    Parse document based on file type.
    """
    ext = filename.lower().split(".")[-1]

    if ext == "pdf":
        return await parse_pdf(document_id)
    elif ext in ["docx", "doc"]:
        return await parse_docx(document_id)
    elif ext in ["xlsx", "xls"]:
        return await parse_spreadsheet(document_id)
    elif ext in ["txt", "md"]:
        return await parse_text(document_id)
    else:
        return f"Unsupported document type: {ext}"


async def parse_pdf(document_id: str) -> str:
    """
    Parse PDF using PyMuPDF.
    """
    try:
        import fitz
        from app.config import settings

        file_path = f"{settings.UPLOAD_DIR}/{document_id}.pdf"
        doc = fitz.open(file_path)

        text_parts = []
        for page_num, page in enumerate(doc):
            text = page.get_text()
            text_parts.append(f"--- Page {page_num + 1} ---\n{text}")

        doc.close()
        return "\n\n".join(text_parts)

    except Exception as e:
        logger.error(f"PDF parsing error: {e}")
        return ""


async def parse_docx(document_id: str) -> str:
    """
    Parse DOCX using python-docx.
    """
    try:
        from docx import Document
        from app.config import settings

        file_path = f"{settings.UPLOAD_DIR}/{document_id}.docx"
        doc = Document(file_path)

        text_parts = []
        for para in doc.paragraphs:
            if para.text.strip():
                text_parts.append(para.text)

        return "\n\n".join(text_parts)

    except Exception as e:
        logger.error(f"DOCX parsing error: {e}")
        return ""


async def parse_spreadsheet(document_id: str) -> str:
    """
    Parse spreadsheet (XLSX) using openpyxl.
    """
    try:
        from openpyxl import load_workbook
        from app.config import settings

        file_path = f"{settings.UPLOAD_DIR}/{document_id}.xlsx"
        wb = load_workbook(file_path, data_only=True)

        text_parts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")
            for row in sheet.iter_rows(max_row=100, values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text_parts.append(row_text)

        return "\n".join(text_parts)

    except Exception as e:
        logger.error(f"Spreadsheet parsing error: {e}")
        return ""


async def parse_text(document_id: str) -> str:
    """
    Parse plain text file.
    """
    try:
        from app.config import settings
        import os

        for ext in ["txt", "md"]:
            file_path = f"{settings.UPLOAD_DIR}/{document_id}.{ext}"
            if os.path.exists(file_path):
                with open(file_path, "r", encoding="utf-8") as f:
                    return f.read()

        return ""

    except Exception as e:
        logger.error(f"Text parsing error: {e}")
        return ""


async def detect_scanned_pages(document_id: str) -> bool:
    """
    Detect if a PDF contains scanned pages (mostly images).
    Uses simple heuristics - in production, use more sophisticated detection.
    """
    try:
        import fitz
        from app.config import settings

        file_path = f"{settings.UPLOAD_DIR}/{document_id}.pdf"
        doc = fitz.open(file_path)

        scanned_count = 0
        for page in doc:
            text = page.get_text().strip()
            if len(text) < 50:
                scanned_count += 1

        doc.close()
        return scanned_count > len(doc) * 0.5

    except Exception:
        return False


async def perform_ocr(document_id: str) -> str:
    """
    Perform OCR using PaddleOCR.
    """
    try:
        from paddleocr import PaddleOCR
        import fitz
        from app.config import settings
        from PIL import Image
        import io

        file_path = f"{settings.UPLOAD_DIR}/{document_id}.pdf"
        doc = fitz.open(file_path)

        ocr = PaddleOCR(use_angle_cls=True, lang="en")
        all_text = []

        for page_num, page in enumerate(doc):
            pix = page.get_pixmap(dpi=200)
            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))

            result = ocr.ocr(img, cls=True)
            page_text = []
            for line in result[0] if result else []:
                page_text.append(line[1][0])
            all_text.append(f"--- Page {page_num + 1} ---\n" + "\n".join(page_text))

        doc.close()
        return "\n\n".join(all_text)

    except Exception as e:
        logger.error(f"OCR error: {e}")
        return ""


async def chunk_text(text: str, doc, chunk_size: int = 512, overlap: int = 64) -> List[Dict[str, Any]]:
    """
    Create structure-aware chunks from text.
    """
    chunks = []

    lines = text.split("\n")
    current_chunk = []
    current_size = 0

    for line in lines:
        line_size = len(line.split())

        if current_size + line_size > chunk_size and current_chunk:
            chunk_text = " ".join(current_chunk)

            chunks.append({
                "chunk_id": str(uuid.uuid4()),
                "text": chunk_text,
                "metadata": {
                    "document_id": doc.id,
                    "filename": doc.filename,
                    "chunk_index": len(chunks),
                },
            })

            overlap_lines = current_chunk[-overlap // 10:] if len(current_chunk) > overlap // 10 else []
            current_chunk = overlap_lines + [line]
            current_size = sum(len(l.split()) for l in current_chunk)
        else:
            current_chunk.append(line)
            current_size += line_size

    if current_chunk:
        chunks.append({
            "chunk_id": str(uuid.uuid4()),
            "text": " ".join(current_chunk),
            "metadata": {
                "document_id": doc.id,
                "filename": doc.filename,
                "chunk_index": len(chunks),
            },
        })

    return chunks


async def generate_embeddings(document_id: str, chunks: List[Dict[str, Any]]):
    """
    Generate embeddings for document chunks.
    Uses mock embeddings for now - replace with actual model.
    """
    from app.rag.dense import dense_retriever

    texts = [chunk["text"] for chunk in chunks]
    embeddings = await dense_retriever.embed(texts)

    for chunk, embedding in zip(chunks, embeddings):
        chunk["embedding"] = embedding


async def index_chunks(document_id: str, chunks: List[Dict[str, Any]]):
    """
    Index chunks in Qdrant.
    """
    from app.storage.qdrant import insert_chunk

    for chunk in chunks:
        await insert_chunk(
            document_id=document_id,
            chunk_id=chunk["chunk_id"],
            text=chunk["text"],
            vector=chunk.get("embedding", [0] * 1024),
            metadata=chunk.get("metadata", {}),
        )
