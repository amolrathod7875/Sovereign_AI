import logging
from typing import Dict, Any, List, Optional

logger = logging.getLogger(__name__)


async def parse_document(document_id: str, filename: str) -> str:
    """
    Parse document based on file type.
    """
    ext = filename.lower().split(".")[-1]

    if ext == "pdf":
        return await parse_pdf(document_id)
    elif ext in ["docx", "doc"]:
        return await parse_docx(document_id)
    elif ext in ["xlsx", "xls"]:
        return await parse_spreadsheet(document_id)
    elif ext in ["txt", "md"]:
        return await parse_text(document_id)
    else:
        return f"Unsupported document type: {ext}"


async def parse_pdf(document_id: str) -> str:
    """
    Parse PDF using PyMuPDF.
    """
    try:
        import fitz
        from app.config import settings

        file_path = f"{settings.UPLOAD_DIR}/{document_id}.pdf"
        doc = fitz.open(file_path)

        text_parts = []
        for page_num, page in enumerate(doc):
            text = page.get_text()
            text_parts.append(f"--- Page {page_num + 1} ---\n{text}")

        doc.close()
        return "\n\n".join(text_parts)

    except Exception as e:
        logger.error(f"PDF parsing error: {e}")
        return ""


async def parse_docx(document_id: str) -> str:
    """
    Parse DOCX using python-docx.
    """
    try:
        from docx import Document
        from app.config import settings

        file_path = f"{settings.UPLOAD_DIR}/{document_id}.docx"
        doc = Document(file_path)

        text_parts = []
        for para in doc.paragraphs:
            if para.text.strip():
                text_parts.append(para.text)

        return "\n\n".join(text_parts)

    except Exception as e:
        logger.error(f"DOCX parsing error: {e}")
        return ""


async def parse_spreadsheet(document_id: str) -> str:
    """
    Parse spreadsheet (XLSX) using openpyxl.
    """
    try:
        from openpyxl import load_workbook
        from app.config import settings

        file_path = f"{settings.UPLOAD_DIR}/{document_id}.xlsx"
        wb = load_workbook(file_path, data_only=True)

        text_parts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")
            for row in sheet.iter_rows(max_row=100, values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text_parts.append(row_text)

        return "\n".join(text_parts)

    except Exception as e:
        logger.error(f"Spreadsheet parsing error: {e}")
        return ""


async def parse_text(document_id: str) -> str:
    """
    Parse plain text file.
    """
    try:
        from app.config import settings
        import os

        for ext in ["txt", "md"]:
            file_path = f"{settings.UPLOAD_DIR}/{document_id}.{ext}"
            if os.path.exists(file_path):
                with open(file_path, "r", encoding="utf-8") as f:
                    return f.read()

        return ""

    except Exception as e:
        logger.error(f"Text parsing error: {e}")
        return ""
