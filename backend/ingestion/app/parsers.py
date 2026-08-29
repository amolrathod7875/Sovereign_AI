import json
import csv
import mimetypes
from pathlib import Path
from typing import Tuple, Dict, Any
import pypdf
import docx
import openpyxl

def detect_file_info(filepath: Path) -> Tuple[str, str]:
    """
    Simulates Apache Tika's file type detection and basic metadata extraction.
    Returns (mime_type, parser_to_use).
    """
    mime_type, _ = mimetypes.guess_type(filepath)
    if not mime_type:
        mime_type = "application/octet-stream"
    
    ext = filepath.suffix.lower()
    
    if ext == ".pdf":
        return "application/pdf", "docling_simulated"
    elif ext in [".docx"]:
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document", "unstructured_simulated"
    elif ext in [".xlsx"]:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "unstructured_simulated"
    elif ext == ".csv":
        return "text/csv", "pandas_simulated"
    elif ext == ".json":
        return "application/json", "json_native"
    elif ext == ".txt":
        return "text/plain", "unstructured_simulated"
    elif ext in [".png", ".jpg", ".jpeg"]:
        return mime_type, "tesseract_ocr_simulated"
    
    return mime_type, "unknown"

def parse_pdf(filepath: Path) -> Dict[str, Any]:
    """Simulates Docling processing a PDF."""
    pages = []
    with open(filepath, "rb") as f:
        reader = pypdf.PdfReader(f)
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            pages.append({
                "page_number": i + 1,
                "elements": [
                    {"type": "paragraph", "text": text.strip()}
                ]
            })
    return {"pages": pages}

def parse_docx(filepath: Path) -> Dict[str, Any]:
    """Simulates Unstructured processing a DOCX."""
    doc = docx.Document(filepath)
    pages = []
    elements = []
    for para in doc.paragraphs:
        if para.text.strip():
            elements.append({"type": "paragraph", "text": para.text.strip()})
    pages.append({"page_number": 1, "elements": elements})
    return {"pages": pages}

def parse_xlsx(filepath: Path) -> Dict[str, Any]:
    """Simulates Unstructured / Pandas processing an XLSX."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    records = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            if any(row):
                records.append({"sheet": sheet, "row": [str(cell) for cell in row if cell is not None]})
    return {"records": records}

def parse_csv(filepath: Path) -> Dict[str, Any]:
    """Processes a CSV file."""
    records = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append(row)
    return {"records": records}

def parse_json(filepath: Path) -> Dict[str, Any]:
    """Processes a JSON file."""
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return {"records": data}
    return {"records": [data]}

def parse_txt(filepath: Path) -> Dict[str, Any]:
    """Processes a text file."""
    with open(filepath, "r", encoding="utf-8") as f:
        text = f.read()
    return {
        "pages": [
            {
                "page_number": 1,
                "elements": [{"type": "paragraph", "text": text}]
            }
        ]
    }

def parse_image(filepath: Path) -> Dict[str, Any]:
    """Simulates OCR on an image."""
    return {
        "pages": [
            {
                "page_number": 1,
                "elements": [{"type": "paragraph", "text": "[SIMULATED OCR TEXT EXTRACTED FROM IMAGE]"}]
            }
        ]
    }

def process_file(filepath: Path, parser_name: str) -> Dict[str, Any]:
    if parser_name == "docling_simulated":
        return parse_pdf(filepath)
    elif parser_name == "unstructured_simulated" and filepath.suffix.lower() == ".docx":
        return parse_docx(filepath)
    elif parser_name == "unstructured_simulated" and filepath.suffix.lower() == ".xlsx":
        return parse_xlsx(filepath)
    elif parser_name == "pandas_simulated" or filepath.suffix.lower() == ".csv":
        return parse_csv(filepath)
    elif parser_name == "json_native":
        return parse_json(filepath)
    elif parser_name == "unstructured_simulated" and filepath.suffix.lower() == ".txt":
        return parse_txt(filepath)
    elif parser_name == "tesseract_ocr_simulated":
        return parse_image(filepath)
    else:
        raise ValueError(f"Unsupported parser or file type: {parser_name} for {filepath}")
